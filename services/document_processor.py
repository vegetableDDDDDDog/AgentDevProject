"""
异步文档处理服务。

支持文件上传、解析、分块、向量化的异步处理流程。
"""

import os
import shutil
import logging
from typing import Optional
from datetime import datetime, timezone
from fastapi import BackgroundTasks
from langchain_text_splitters import RecursiveCharacterTextSplitter

from services.database import (
    KnowledgeBase, Document, DocumentProcessingTask,
    get_db, SessionLocal
)
from services.document_type_detector import DocumentTypeDetector
from services.ocr_service import OCRServiceFactory
from api.config import settings

logger = logging.getLogger(__name__)


class DocumentProcessor:
    """异步文档处理器"""

    def __init__(self, db_session):
        """
        初始化文档处理器。

        Args:
            db_session: SQLAlchemy 数据库会话
        """
        self.db = db_session
        self.type_detector = DocumentTypeDetector(ocr_threshold=10)

        # 创建 OCR 服务（使用 Mock 模式，避免 Tesseract 依赖）
        self.ocr_service = OCRServiceFactory.create_service({
            'provider': 'mock'  # 可配置为 'tesseract'
        })
        logger.info(f"DocumentProcessor 初始化,OCR: {self.ocr_service.name}")

    async def upload_document_async(
        self,
        file,
        knowledge_base_id: str,
        tenant_id: str,
        background_tasks: BackgroundTasks
    ) -> str:
        """
        异步上传文档。

        Args:
            file: FastAPI UploadFile 对象
            knowledge_base_id: 知识库 ID
            tenant_id: 租户 ID
            background_tasks: FastAPI BackgroundTasks

        Returns:
            task_id: 任务 ID(立即返回)
        """
        # 1. 保存文件到本地
        file_path = await self._save_file(file, tenant_id)

        # 2. 创建文档记录
        document = Document(
            knowledge_base_id=knowledge_base_id,
            tenant_id=tenant_id,
            filename=file.filename,
            file_type=file.filename.split('.')[-1].lower(),
            file_size=os.path.getsize(file_path),
            file_path=file_path,
            upload_status="processing"
        )
        self.db.add(document)
        self.db.commit()

        # 3. 创建处理任务
        task = DocumentProcessingTask(
            document_id=document.id,
            tenant_id=tenant_id,
            status="processing",
            progress=0,
            current_step="任务已创建"
        )
        self.db.add(task)
        self.db.commit()

        # 4. 提交后台任务
        background_tasks.add_task(
            self._process_document_background,
            task.id,
            document.id,
            file_path
        )

        logger.info(f"文档 {file.filename} 已提交处理,任务 ID: {task.id}")
        return task.id

    async def _process_document_background(
        self,
        task_id: str,
        document_id: str,
        file_path: str
    ):
        """
        后台异步处理文档。

        Args:
            task_id: 任务 ID
            document_id: 文档 ID
            file_path: 文件路径
        """
        db = SessionLocal()

        try:
            task = db.query(DocumentProcessingTask).get(task_id)
            document = db.query(Document).get(document_id)
            kb = db.query(KnowledgeBase).get(document.knowledge_base_id)

            # Step 1: 文件类型检测 (10-15%)
            await self._update_progress(db, task, 10, "正在分析文件类型...")
            detection_result = self.type_detector.detect_document_type(
                file_path, kb
            )

            logger.info(
                f"文件类型检测: {detection_result['type']}, "
                f"需要 OCR: {detection_result['needs_ocr']}"
            )

            # Step 2: 解析文本 (15-40%)
            if detection_result['needs_ocr']:
                await self._update_progress(db, task, 15, "检测到扫描件,正在 OCR 识别...")
                text = await self.ocr_service.extract_text(file_path)
                document.ocr_used = True
            else:
                await self._update_progress(db, task, 15, "正在解析文档...")
                text = await self._parse_file(file_path)

            await self._update_progress(db, task, 40, "文档解析完成")

            # Step 3: 分块 (40-60%)
            await self._update_progress(db, task, 40, "正在分块...")
            chunks = self._split_text(text, kb)

            # Step 4: 向量化 (60-90%)
            await self._update_progress(db, task, 60, "正在向量化...")
            await self._vectorize_and_store(chunks, kb, document)

            # 完成
            await self._update_progress(db, task, 100, "处理完成")
            self._mark_task_completed(db, task, document, len(chunks))

            # 完成
            await self._update_progress(db, task, 100, "处理完成")
            self._mark_task_completed(db, task, document, len(chunks))

            logger.info(f"文档 {document.filename} 处理完成")

        except Exception as e:
            logger.error(f"文档处理失败: {e}", exc_info=True)
            self._mark_task_failed(db, task, str(e))

        finally:
            db.close()

    async def _update_progress(
        self,
        db,
        task: DocumentProcessingTask,
        progress: int,
        current_step: str
    ):
        """更新任务进度"""
        task.progress = progress
        task.current_step = current_step
        task.updated_at = datetime.now(timezone.utc)
        db.commit()

    def _mark_task_completed(
        self,
        db,
        task: DocumentProcessingTask,
        document: Document,
        chunk_count: int
    ):
        """标记任务完成"""
        task.status = "completed"
        task.progress = 100
        task.completed_at = datetime.now(timezone.utc)

        document.upload_status = "ready"
        document.processed_at = datetime.now(timezone.utc)
        document.chunk_count = chunk_count

        # 更新知识库统计
        kb = db.query(KnowledgeBase).get(document.knowledge_base_id)
        kb.document_count += 1
        kb.total_chunks += chunk_count
        kb.updated_at = datetime.now(timezone.utc)

        db.commit()

    def _mark_task_failed(self, db, task: DocumentProcessingTask, error: str):
        """标记任务失败"""
        task.status = "failed"
        task.error_message = error
        task.completed_at = datetime.now(timezone.utc)

        document = db.query(Document).get(task.document_id)
        document.upload_status = "failed"

        db.commit()

    async def _save_file(self, file, tenant_id: str) -> str:
        """保存文件到本地"""
        upload_dir = f"data/knowledge_bases/{tenant_id}"
        os.makedirs(upload_dir, exist_ok=True)

        file_path = f"{upload_dir}/{file.filename}"
        with open(file_path, "wb") as f:
            shutil.copyfileobj(file.file, f)

        return file_path

    async def _parse_file(self, file_path: str) -> str:
        """解析文件(文本/Markdown/Excel)"""
        ext = file_path.lower().split('.')[-1]

        if ext in ['md', 'txt']:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()

        elif ext == 'pdf':
            import pypdf
            pdf_reader = pypdf.PdfReader(file_path)
            return '\n'.join([page.extract_text() for page in pdf_reader.pages])

        elif ext in ['xlsx', 'xls']:
            import openpyxl
            wb = openpyxl.load_workbook(file_path)
            sheets = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append('\t'.join([str(cell) for cell in row if cell is not None]))
                sheets.append(f"## {sheet_name}\n" + '\n'.join(data))
            return '\n\n'.join(sheets)

        else:
            raise ValueError(f"不支持的文件类型: {ext}")

    def _split_text(self, text: str, kb: KnowledgeBase) -> list:
        """分块"""
        splitter = RecursiveCharacterTextSplitter(
            chunk_size=kb.chunk_size,
            chunk_overlap=kb.chunk_overlap,
            separators=["\n\n##", "\n\n", "\n", "。", "，", " ", ""]
        )
        return splitter.split_text(text)

    async def _vectorize_and_store(
        self,
        chunks: list,
        kb: KnowledgeBase,
        document: Document
    ):
        """
        向量化并存储到 ChromaDB。

        Args:
            chunks: 文本分块列表
            kb: 知识库
            document: 文档 ORM 对象
        """
        try:
            from langchain_core.documents import Document
            from langchain_community.vectorstores import Chroma
            from services.embeddings_service import get_embeddings_service
            import uuid

            embeddings = get_embeddings_service()

            # 创建 LangChain Document 对象
            documents = []
            for i, chunk in enumerate(chunks):
                doc = Document(
                    page_content=chunk,
                    metadata={
                        "document_id": document.id,
                        "chunk_id": f"{document.id}_{i}",
                        "filename": document.filename,
                        "tenant_id": kb.tenant_id
                    }
                )
                documents.append(doc)

            # 创建 ChromaDB 持久目录
            persist_dir = f"data/chroma/{kb.tenant_id}"
            os.makedirs(persist_dir, exist_ok=True)

            # 存储到 ChromaDB
            vector_store = Chroma(
                collection_name=kb.collection_name,
                embedding_function=embeddings.langchain_embeddings,
                persist_directory=persist_dir
            )
            vector_store.add_documents(documents)

            logger.info(
                f"向量化和存储完成: {len(documents)} 个 chunks "
                f"→ collection: {kb.collection_name}"
            )

        except Exception as e:
            logger.error(f"向量化失败: {e}", exc_info=True)
            raise
